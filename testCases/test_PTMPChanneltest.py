#!/usr/bin/env python3
"""
PTMP Channel Connectivity Test — pytest version
=================================================
All parameters come from conftest.py fixtures / CLI options.
Run via Jenkins:
    python3 -m pytest -v -s testCases/test_PTMPChannelTest.py::test_PTMPChannelConnectivity \
        --local-ip        192.168.1.111               \
        --remote-ips      "192.168.1.110,192.168.1.111" \
        --radio           Radio1                       \
        --bandwidth       HT80                         \
        --country         "5GHz"                       \
        --channels        "36,40,100,104,144,169,171"  \
        --snr-settle-delay  45                         \
        --monitor-duration  60                         \
        --monitor-interval  10                         \
        --connect-timeout   180                        \
        --bw-settle-wait    60                         \
        --chan-settle-wait  10                         \
        --username          root                       \
        --password          admin                      \
        --snmp-community    ubr@rw123                  \
        --output-prefix     ptmp_channel_test
"""

import re
import time
import platform
import subprocess
import csv
import json
from datetime import datetime

try:
    import paramiko
    _paramiko_ok = True
except ImportError:
    print("WARNING: 'paramiko' not installed — SSH channel-list fetch disabled.")
    _paramiko_ok = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    _excel_ok = True
except ImportError:
    print("WARNING: 'openpyxl' not installed — Excel report disabled.")
    _excel_ok = False


# ═══════════════════════════════════════════════════════════
#  ANSI colour helpers
# ═══════════════════════════════════════════════════════════
class C:
    HDR   = '\033[95m'
    BLUE  = '\033[94m'
    GREEN = '\033[92m'
    WARN  = '\033[93m'
    FAIL  = '\033[91m'
    END   = '\033[0m'
    BOLD  = '\033[1m'


# ═══════════════════════════════════════════════════════════
#  Country code map
# ═══════════════════════════════════════════════════════════
COUNTRY_CODES = {
    "US 5GHz All":     5012,
    "US 5GHz Non-DFS": 5011,
    "Europe":          276,
    "Canada":          124,
    "5GHz":            5019,
    "India":           356,
}


# ═══════════════════════════════════════════════════════════
#  Low-level helpers
# ═══════════════════════════════════════════════════════════
def ping(host):
    param = "-n" if platform.system().lower() == "windows" else "-c"
    return subprocess.call(
        ["ping", param, "1", "-W", "2", host],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
    ) == 0


def _run(cmd):
    try:
        return subprocess.check_output(
            cmd, shell=True, stderr=subprocess.DEVNULL
        ).decode("utf-8", errors="replace").strip()
    except Exception:
        return None


def snmp_get(ip, oid, community):
    out = _run(f"snmpget -v 2c -c {community} {ip} {oid}")
    if out is None:
        return "Err"
    m = re.search(r'(?:INTEGER|STRING|Gauge32|Counter32|IpAddress|Timeticks):\s*(.+)', out)
    return m.group(1).replace('"', '').strip() if m else out.split(":")[-1].strip()


def snmp_set(ip, oid, type_char, value, community):
    try:
        subprocess.check_output(
            f"snmpset -v 2c -c {community} {ip} {oid} {type_char} {value}",
            shell=True, stderr=subprocess.STDOUT
        )
        return True
    except subprocess.CalledProcessError as e:
        print(f"{C.FAIL}snmpset failed ({oid}={value}): "
              f"{e.output.decode(errors='replace').strip()}{C.END}", flush=True)
        return False


def apply_config(ip, community):
    snmp_set(ip, ".1.3.6.1.4.1.52619.1.2.1.1.0", "i", "1", community)


def parse_uptime_seconds(s):
    try:
        if not s or s in ("Err", "-"):
            return -1
        if "(" in s:
            s = s.split("(")[-1].split(")")[0]
        parts = [int(x) for x in s.split(":")]
        parts.reverse()
        mults = [1, 60, 3600, 86400]
        return sum(v * mults[i] for i, v in enumerate(parts) if i < 4)
    except Exception:
        return -1


def wait_for_ping(ip, timeout=200, interval=10):
    elapsed = 0
    while elapsed < timeout:
        if ping(ip):
            return True
        time.sleep(interval)
        elapsed += interval
    return False


def wait_for_all_links(ip_list, timeout):
    elapsed = 0
    while elapsed < timeout:
        if all(ping(ip) for ip in ip_list):
            return True
        time.sleep(5)
        elapsed += 5
    return False


# ═══════════════════════════════════════════════════════════
#  Radio OID helpers
# ═══════════════════════════════════════════════════════════
def radio_oid(radio):   return "2" if radio == "Radio1" else "3"
def radio_index(radio): return "1" if radio == "Radio1" else "2"


# ═══════════════════════════════════════════════════════════
#  BSU configuration helpers
# ═══════════════════════════════════════════════════════════
def set_bandwidth(ip, radio, bw, community, settle_wait):
    oid = f".1.3.6.1.4.1.52619.1.1.1.1.1.7.{radio_oid(radio)}"
    print(f"\n{C.HDR}[BW] Setting bandwidth -> {bw}{C.END}", flush=True)
    for attempt in range(1, 4):
        if snmp_set(ip, oid, "s", bw, community):
            time.sleep(2)
            apply_config(ip, community)
            print(f"     Waiting {settle_wait}s for bandwidth to apply ...", flush=True)
            time.sleep(settle_wait)
            return True
        print(f"     Attempt {attempt}/3 failed. Retrying ...", flush=True)
        time.sleep(5)
    print(f"{C.FAIL}CRITICAL: Failed to set bandwidth after 3 attempts!{C.END}", flush=True)
    return False


def set_channel(ip, radio, channel, community, settle_wait):
    oid = f".1.3.6.1.4.1.52619.1.1.1.1.1.9.{radio_oid(radio)}"
    print(f"  {C.BLUE}[CH] Setting channel -> {channel}{C.END}", flush=True)
    for attempt in range(1, 4):
        if snmp_set(ip, oid, "i", channel, community):
            time.sleep(2)
            apply_config(ip, community)
            print(f"       Waiting {settle_wait}s for channel to apply ...", flush=True)
            time.sleep(settle_wait)
            return True
        print(f"       Attempt {attempt}/3 failed. Retrying ...", flush=True)
        time.sleep(5)
    print(f"{C.FAIL}CRITICAL: Failed to set channel after 3 attempts!{C.END}", flush=True)
    return False


def verify_bsu_operation(ip, radio, expected_bw, expected_ch, community):
    roid   = radio_oid(radio)
    op_bw  = snmp_get(ip, f".1.3.6.1.4.1.52619.1.1.1.1.1.51.{roid}", community)
    op_ch  = snmp_get(ip, f".1.3.6.1.4.1.52619.1.1.1.1.1.23.{roid}", community)
    bw_ok  = expected_bw.replace("HT", "") in op_bw or op_bw in expected_bw
    ch_ok  = str(op_ch).strip() == str(expected_ch).strip()
    bw_tag = f"{C.GREEN}OK{C.END}"    if bw_ok else f"{C.WARN}MISMATCH{C.END}"
    ch_tag = f"{C.GREEN}MATCH{C.END}" if ch_ok else f"{C.FAIL}MISMATCH{C.END}"
    print(f"  BSU Op BW : {op_bw:<10} (target {expected_bw}) [{bw_tag}]", flush=True)
    print(f"  BSU Op CH : {op_ch:<10} (target {expected_ch}) [{ch_tag}]", flush=True)
    return ch_ok, bw_ok, op_bw, op_ch


# ═══════════════════════════════════════════════════════════
#  SSH channel-list fetch
# ═══════════════════════════════════════════════════════════
def fetch_channel_list_ssh(ip, radio, country_code, bw, ssh_user, ssh_pass):
    if not _paramiko_ok:
        return []
    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(ip, username=ssh_user, password=ssh_pass, timeout=15)
        rind = radio_index(radio)
        cmd  = f"/usr/sbin/kwn_get_supp_chan.sh {rind} {country_code} {bw}"
        _, stdout, _ = ssh.exec_command(cmd)
        nums = re.findall(r'\b\d+\b', stdout.read().decode("utf-8", errors="replace"))
        ssh.close()
        return nums[::2]
    except Exception as e:
        print(f"{C.FAIL}SSH channel-list fetch failed: {e}{C.END}", flush=True)
        return []


# ═══════════════════════════════════════════════════════════
#  ONE-SHOT SNR / link-stats fetch  (called once per channel)
# ═══════════════════════════════════════════════════════════
def fetch_all_link_stats_once(local_ip, radio, community):
    roid  = radio_oid(radio)
    found = {}

    def _int_val(suffix, idx):
        out = _run(f"snmpget -v 2c -c {community} {local_ip} "
                   f".1.3.6.1.4.1.52619.1.3.3.1.{suffix}.{roid}.{idx}")
        if not out:
            return "-"
        m = re.search(r'INTEGER:\s*(\d+)', out)
        return m.group(1) if m else "-"

    def _uptime_val(idx):
        out = _run(f"snmpget -v 2c -c {community} {local_ip} "
                   f".1.3.6.1.4.1.52619.1.3.3.1.52.{roid}.{idx}")
        if not out:
            return "Err"
        raw = out.split("=")[-1].strip()
        raw = re.sub(r'^(STRING|Timeticks|INTEGER):\s*', '', raw)
        m   = re.search(r'\((.*?)\)', raw)
        if m:
            raw = m.group(1)
        return raw.replace('"', '').replace('(', '').replace(')', '').strip()

    for idx in range(1, 33):
        out_ip = _run(f"snmpget -v 2c -c {community} {local_ip} "
                      f".1.3.6.1.4.1.52619.1.3.3.1.4.{roid}.{idx}")
        if not out_ip or "No Such Instance" in out_ip:
            continue
        m_ip = re.search(r'IpAddress:\s*([\d.]+)', out_ip)
        ip   = m_ip.group(1) if m_ip else "-"
        if ip in ("-", "0.0.0.0"):
            continue

        found[ip] = {
            "l_snr_a1": _int_val("13", idx),
            "l_snr_a2": _int_val("14", idx),
            "r_snr_a1": _int_val("15", idx),
            "r_snr_a2": _int_val("16", idx),
            "tx_rate":  _int_val("10", idx),
            "rx_rate":  _int_val("9",  idx),
            "noise":    _int_val("17", idx),
            "uptime":   _uptime_val(idx),
        }
        d = found[ip]
        print(
            f"  {C.GREEN}[STATS] {ip:<16} "
            f"L_SNR:{d['l_snr_a1']}/{d['l_snr_a2']}  "
            f"R_SNR:{d['r_snr_a1']}/{d['r_snr_a2']}  "
            f"TX:{d['tx_rate']}  RX:{d['rx_rate']}  "
            f"Noise:{d['noise']}  Up:{d['uptime']}{C.END}", flush=True
        )
    return found


# ═══════════════════════════════════════════════════════════
#  Fast uptime-only poll  (used during monitoring loop)
# ═══════════════════════════════════════════════════════════
def fetch_uptime_only(local_ip, radio, community, su_indices):
    roid   = radio_oid(radio)
    result = {}
    for ip, idx in su_indices.items():
        out = _run(f"snmpget -v 2c -c {community} {local_ip} "
                   f".1.3.6.1.4.1.52619.1.3.3.1.52.{roid}.{idx}")
        if not out:
            result[ip] = "Err"
            continue
        raw = out.split("=")[-1].strip()
        raw = re.sub(r'^(STRING|Timeticks|INTEGER):\s*', '', raw)
        m   = re.search(r'\((.*?)\)', raw)
        if m:
            raw = m.group(1)
        result[ip] = raw.replace('"', '').replace('(', '').replace(')', '').strip()
    return result


def discover_su_indices(local_ip, radio, community, remote_ips_set):
    roid    = radio_oid(radio)
    indices = {}
    for idx in range(1, 33):
        out = _run(f"snmpget -v 2c -c {community} {local_ip} "
                   f".1.3.6.1.4.1.52619.1.3.3.1.4.{roid}.{idx}")
        if not out or "No Such Instance" in out:
            continue
        m  = re.search(r'IpAddress:\s*([\d.]+)', out)
        ip = m.group(1) if m else "-"
        if ip in remote_ips_set:
            indices[ip] = idx
    return indices


# ═══════════════════════════════════════════════════════════
#  Stability monitoring loop  (uptime-reform check only)
# ═══════════════════════════════════════════════════════════
def monitor_stability(local_ip, remote_ips, radio, channel, bw,
                      community, duration, interval,
                      su_indices, txt_file, csv_writer,
                      snr_snapshot, timestamp_of_snap):
    issues         = []
    prev_uptime    = {}
    elapsed        = 0
    zero_snr_abort = False

    for ip in remote_ips:
        d = snr_snapshot.get(ip)
        if d and any(d.get(k, "1") == "0"
                     for k in ("l_snr_a1", "l_snr_a2", "r_snr_a1", "r_snr_a2")):
            issues.append(f"{ip}: Zero SNR (snapshot)")
            zero_snr_abort = True

    if zero_snr_abort:
        print(f"{C.FAIL}  CRITICAL: Zero SNR in snapshot — aborting stability window.{C.END}", flush=True)
        return issues

    print(f"\n  {C.BOLD}Stability monitoring {duration}s (poll every {interval}s) ...{C.END}", flush=True)

    while elapsed < duration:
        loop_start  = time.time()
        ts          = datetime.now().strftime("%H:%M:%S")
        uptime_map  = fetch_uptime_only(local_ip, radio, community, su_indices)

        for ip in remote_ips:
            uptime_str = uptime_map.get(ip, "Err")
            cur_sec    = parse_uptime_seconds(uptime_str)
            prev_sec   = prev_uptime.get(ip, -1)
            note       = ""
            status     = "PASS"
            color      = C.GREEN

            if prev_sec != -1 and cur_sec != -1 and cur_sec < prev_sec:
                status = "FAIL"; note = "LINK REFORMED"; color = C.FAIL
                issues.append(f"{ip}: LINK REFORMED")

            if cur_sec != -1:
                prev_uptime[ip] = cur_sec

            d = snr_snapshot.get(ip, {})
            print(
                f"  {color}[{ts}] {ip:<16} "
                f"L:{d.get('l_snr_a1','-')}/{d.get('l_snr_a2','-')}  "
                f"R:{d.get('r_snr_a1','-')}/{d.get('r_snr_a2','-')}  "
                f"TX:{d.get('tx_rate','-')}  RX:{d.get('rx_rate','-')}  "
                f"Up:{uptime_str}  {status} {note}{C.END}", flush=True
            )
            txt_file.write(
                f"[{ts}] CH:{channel} BW:{bw} IP:{ip} "
                f"L:{d.get('l_snr_a1','-')}/{d.get('l_snr_a2','-')} "
                f"R:{d.get('r_snr_a1','-')}/{d.get('r_snr_a2','-')} "
                f"Up:{uptime_str} {status} {note}\n"
            )
            csv_writer.writerow([
                ts, channel, bw, ip,
                f"{d.get('l_snr_a1','-')}/{d.get('l_snr_a2','-')}",
                f"{d.get('r_snr_a1','-')}/{d.get('r_snr_a2','-')}",
                f"{d.get('tx_rate','-')}/{d.get('rx_rate','-')}",
                d.get("noise", "-"), uptime_str, status, note,
                f"(SNR snapshot @ {timestamp_of_snap})"
            ])

        sleep = max(0, interval - (time.time() - loop_start))
        time.sleep(sleep)
        elapsed += interval
        print(f"  Elapsed {elapsed}/{duration}s", flush=True)

    return issues


# ═══════════════════════════════════════════════════════════
#  Excel + JSON report helpers
# ═══════════════════════════════════════════════════════════
def generate_excel(results, xlsx_path):
    if not _excel_ok:
        return
    print(f"\n{C.HDR}Generating Excel: {xlsx_path}{C.END}", flush=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Channel Summary"
    headers = ["Channel", "Bandwidth", "Country", "Status",
               "L SNR A1/A2", "R SNR A1/A2", "TX Rate", "RX Rate",
               "Noise", "Uptime (snapshot)", "Notes"]
    ws.append(headers)
    bold   = Font(bold=True)
    c_pass = PatternFill(start_color="00CC44", end_color="00CC44", fill_type="solid")
    c_fail = PatternFill(start_color="FF3333", end_color="FF3333", fill_type="solid")
    c_warn = PatternFill(start_color="FFAA00", end_color="FFAA00", fill_type="solid")
    center = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.font = bold; cell.alignment = center
    for row_idx, r in enumerate(results, start=2):
        snap = r.get("snr_snapshot", {})
        def agg(key):
            vals = [snap.get(ip, {}).get(key, "-") for ip in r.get("remote_ips", [])]
            return " | ".join(vals) if vals else "-"
        ws.cell(row=row_idx, column=1,  value=r["channel"])
        ws.cell(row=row_idx, column=2,  value=r["bandwidth"])
        ws.cell(row=row_idx, column=3,  value=r["country"])
        cell_s = ws.cell(row=row_idx, column=4, value=r["status"])
        ws.cell(row=row_idx, column=5,  value=f"{agg('l_snr_a1')}/{agg('l_snr_a2')}")
        ws.cell(row=row_idx, column=6,  value=f"{agg('r_snr_a1')}/{agg('r_snr_a2')}")
        ws.cell(row=row_idx, column=7,  value=agg("tx_rate"))
        ws.cell(row=row_idx, column=8,  value=agg("rx_rate"))
        ws.cell(row=row_idx, column=9,  value=agg("noise"))
        ws.cell(row=row_idx, column=10, value=agg("uptime"))
        ws.cell(row=row_idx, column=11, value=r["notes"])
        cell_s.fill = c_pass if r["status"] == "PASS" else (c_fail if r["status"] == "FAIL" else c_warn)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(cell.value or "")) for cell in col) + 4, 40
        )
    wb.save(xlsx_path)
    print(f"{C.GREEN}Excel saved -> {xlsx_path}{C.END}", flush=True)


def write_json(results, meta, json_path):
    with open(json_path, "w") as f:
        json.dump({"generated": datetime.now().isoformat(), "meta": meta, "results": results}, f, indent=2)
    print(f"{C.GREEN}JSON saved -> {json_path}{C.END}", flush=True)


# ═══════════════════════════════════════════════════════════
#  pytest test function — all config via conftest fixtures
# ═══════════════════════════════════════════════════════════
def test_PTMPChannelConnectivity(
    local_ip,           # --local-ip
    remote_ips,         # --remote-ips  -> list of IPs
    radio,              # --radio
    bandwidth,          # --bandwidth   (comma-sep string, e.g. "HT80" or "HT20,HT80")
    country,            # --country     (comma-sep string)
    channels,           # --channels    -> list of channel strings
    snr_settle_delay,   # --snr-settle-delay  -> int
    monitor_duration,   # --monitor-duration  -> int
    monitor_interval,   # --monitor-interval  -> int
    connect_timeout,    # --connect-timeout   -> int
    bw_settle_wait,     # --bw-settle-wait    -> int
    chan_settle_wait,   # --chan-settle-wait   -> int
    username,           # --username  (SSH user)
    password,           # --password  (SSH pass)
    snmp_community,     # --snmp-community
    output_prefix,      # --output-prefix
):
    # ── Unpack / normalise ────────────────────────────────
    local_ip        = local_ip.strip()
    bandwidths      = [b.strip() for b in bandwidth.split(",")  if b.strip()]
    countries       = [c.strip() for c in country.split(",")    if c.strip()]
    custom_channels = channels      # list already parsed by conftest fixture
    ssh_user        = username
    ssh_pass        = password
    community       = snmp_community
    snr_delay       = snr_settle_delay
    mon_dur         = monitor_duration
    mon_int         = monitor_interval
    conn_to         = connect_timeout
    bw_wait         = bw_settle_wait
    ch_wait         = chan_settle_wait
    prefix          = output_prefix

    txt_path  = f"{prefix}.txt"
    csv_path  = f"{prefix}.csv"
    xlsx_path = f"{prefix}_summary.xlsx"
    json_path = f"{prefix}_results.json"

    all_channel_results = []

    # ── Banner ────────────────────────────────────────────
    print(f"\n{'='*65}", flush=True)
    print(f"  PTMP Channel Connectivity Test  —  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", flush=True)
    print(f"  Local IP       : {local_ip}", flush=True)
    print(f"  Remote SUs     : {remote_ips}", flush=True)
    print(f"  Radio          : {radio}", flush=True)
    print(f"  Bandwidths     : {bandwidths}", flush=True)
    print(f"  Countries      : {countries}", flush=True)
    print(f"  Channels       : {custom_channels if custom_channels else '(auto-fetch)'}", flush=True)
    print(f"  SNR settle     : {snr_delay}s  (one-shot read per channel)", flush=True)
    print(f"  Monitor window : {mon_dur}s  (poll every {mon_int}s)", flush=True)
    print(f"  Connect TO     : {conn_to}s", flush=True)
    print(f"{'='*65}\n", flush=True)

    # ── Init output files ─────────────────────────────────
    with open(txt_path, "w") as f:
        f.write(f"PTMP Channel Connectivity Test — {datetime.now()}\n")
        f.write(f"Local: {local_ip}   Remotes: {remote_ips}   Radio: {radio}\n\n")

    csv_f = open(csv_path, "w", newline="")
    csv_w = csv.writer(csv_f)
    csv_w.writerow([
        "Timestamp", "Channel", "Bandwidth", "Country", "Node_IP",
        "L_SNR_A1/A2", "R_SNR_A1/A2", "Tx/Rx", "Noise",
        "Uptime", "Status", "Notes", "SNR_Read_Info"
    ])

    # ── Main loops: country -> bandwidth -> channel ───────
    for cty in countries:
        country_code = COUNTRY_CODES.get(cty, 5019)
        print(f"\n{C.HDR}  Country: {cty}  (code={country_code}){C.END}", flush=True)

        with open(txt_path, "a") as f:
            f.write(f"\n{'='*60}\nCountry: {cty}\n{'='*60}\n")

        for bw in bandwidths:
            if cty in ("Canada", "US 5GHz Non-DFS") and bw == "HT160":
                print(f"{C.WARN}  Skipping HT160 — not supported in {cty}{C.END}", flush=True)
                continue

            bw_set = "HT40+" if bw == "HT40" else bw
            print(f"\n{C.HDR}  Bandwidth: {bw}  (setting as {bw_set}){C.END}", flush=True)

            wait_for_ping(local_ip)
            set_bandwidth(local_ip, radio, bw_set, community, bw_wait)
            wait_for_ping(local_ip)

            if custom_channels:
                channel_list = custom_channels
                print(f"{C.BLUE}  Custom channels: {channel_list}{C.END}", flush=True)
            else:
                channel_list = fetch_channel_list_ssh(
                    local_ip, radio, country_code, bw_set, ssh_user, ssh_pass
                )
                if not channel_list:
                    print(f"{C.WARN}  No channels fetched — skipping BW {bw}{C.END}", flush=True)
                    continue
                print(f"  Auto-fetched channels: {channel_list}", flush=True)

            for channel in channel_list:
                freq = (int(channel) * 5) + 5000
                print(f"\n{C.BLUE}  > CH {channel} ({freq} MHz)  BW:{bw}  Country:{cty}{C.END}", flush=True)

                wait_for_ping(local_ip)
                set_channel(local_ip, radio, channel, community, ch_wait)

                print(f"  Waiting for {len(remote_ips)} SU(s) to associate ...", flush=True)
                links_up = wait_for_all_links(remote_ips, conn_to)

                if not links_up:
                    msg = f"Link Timeout ({conn_to}s)"
                    print(f"{C.FAIL}  {msg} — skipping channel {channel}{C.END}", flush=True)
                    all_channel_results.append({
                        "channel": channel, "freq_mhz": freq, "bandwidth": bw,
                        "country": cty,     "status": "FAIL", "notes": msg,
                        "snr_snapshot": {}, "remote_ips": remote_ips,
                    })
                    with open(txt_path, "a") as f:
                        f.write(f"\n[FAIL] CH:{channel} BW:{bw} — {msg}\n")
                    csv_w.writerow([
                        datetime.now().strftime("%H:%M:%S"),
                        channel, bw, cty, "ALL",
                        "-/-", "-/-", "-/-", "-", "-", "FAIL", msg, "-"
                    ])
                    continue

                print(f"{C.GREEN}  All SU links formed!{C.END}", flush=True)

                ch_ok, bw_ok, op_bw, op_ch = verify_bsu_operation(
                    local_ip, radio, bw_set, channel, community
                )
                if not ch_ok:
                    print(f"{C.WARN}  WARNING: BSU channel mismatch "
                          f"(expected {channel}, got {op_ch}){C.END}", flush=True)

                # ── ONE-SHOT SNR fetch after settle delay ─────────
                print(f"\n  Waiting {snr_delay}s before reading SNR ...", flush=True)
                time.sleep(snr_delay)

                snap_ts      = datetime.now().strftime("%H:%M:%S")
                print(f"  Reading SNR / link-stats (one-shot @ {snap_ts}) ...", flush=True)
                snr_snapshot = fetch_all_link_stats_once(local_ip, radio, community)

                for ip in remote_ips:
                    d      = snr_snapshot.get(ip, {})
                    status = "PASS"
                    note   = "SNR snapshot"
                    if not d:
                        status = "FAIL"; note = "Missing from SNMP table at snapshot"
                        print(f"{C.FAIL}  [SNAPSHOT] {ip} — not found in BSU SNMP table{C.END}", flush=True)
                    elif any(d.get(k, "1") == "0"
                             for k in ("l_snr_a1", "l_snr_a2", "r_snr_a1", "r_snr_a2")):
                        status = "FAIL"; note = "Zero SNR at snapshot"

                    csv_w.writerow([
                        snap_ts, channel, bw, cty, ip,
                        f"{d.get('l_snr_a1','-')}/{d.get('l_snr_a2','-')}",
                        f"{d.get('r_snr_a1','-')}/{d.get('r_snr_a2','-')}",
                        f"{d.get('tx_rate','-')}/{d.get('rx_rate','-')}",
                        d.get("noise", "-"), d.get("uptime", "-"),
                        status, note, f"One-shot snapshot @ {snap_ts}"
                    ])
                    with open(txt_path, "a") as f:
                        f.write(
                            f"[SNAPSHOT@{snap_ts}] CH:{channel} BW:{bw} IP:{ip} "
                            f"L:{d.get('l_snr_a1','-')}/{d.get('l_snr_a2','-')} "
                            f"R:{d.get('r_snr_a1','-')}/{d.get('r_snr_a2','-')} "
                            f"TX:{d.get('tx_rate','-')} Noise:{d.get('noise','-')} "
                            f"Up:{d.get('uptime','-')}\n"
                        )

                su_indices = discover_su_indices(local_ip, radio, community, set(remote_ips))

                print(f"\n  Starting {mon_dur}s stability window ...", flush=True)
                with open(txt_path, "a") as txt_f:
                    issues = monitor_stability(
                        local_ip, remote_ips, radio, channel, bw,
                        community, mon_dur, mon_int,
                        su_indices, txt_f, csv_w,
                        snr_snapshot, snap_ts
                    )

                unique_issues = list(set(issues))
                status = "PASS" if not unique_issues else "FAIL"
                notes  = "Stable" if not unique_issues else "; ".join(unique_issues[:5])

                all_channel_results.append({
                    "channel":      channel,
                    "freq_mhz":     freq,
                    "bandwidth":    bw,
                    "country":      cty,
                    "status":       status,
                    "notes":        notes,
                    "snr_snapshot": snr_snapshot,
                    "remote_ips":   remote_ips,
                    "op_bw":        op_bw,
                    "op_ch":        str(op_ch),
                    "snap_ts":      snap_ts,
                })

                with open(txt_path, "a") as f:
                    f.write(f"\n[SUMMARY] CH:{channel} BW:{bw} Country:{cty} -> {status} — {notes}\n")

                color = C.GREEN if status == "PASS" else C.FAIL
                print(f"{color}  > CH:{channel} ({freq} MHz) {bw} {cty} -> {status}  ({notes}){C.END}\n",
                      flush=True)

    csv_f.close()

    # ── Reports ───────────────────────────────────────────
    generate_excel(all_channel_results, xlsx_path)
    write_json(
        all_channel_results,
        {"local_ip": local_ip, "remote_ips": remote_ips, "radio": radio,
         "bandwidths": bandwidths, "countries": countries, "snr_delay_s": snr_delay},
        json_path
    )

    # ── Final console summary ─────────────────────────────
    print(f"\n{'='*65}", flush=True)
    print(f"  FINAL SUMMARY", flush=True)
    print(f"{'='*65}", flush=True)
    for r in all_channel_results:
        c = C.GREEN if r["status"] == "PASS" else C.FAIL
        print(f"  {c}CH {r['channel']:>5} ({r['freq_mhz']} MHz)  "
              f"{r['bandwidth']:<6}  {r['country']:<20}  "
              f"{r['status']}  — {r['notes']}{C.END}", flush=True)

    total   = len(all_channel_results)
    passed  = sum(1 for r in all_channel_results if r["status"] == "PASS")
    overall = "PASS" if passed == total and total > 0 else ("PARTIAL" if passed > 0 else "FAIL")
    oc      = C.GREEN if overall == "PASS" else (C.WARN if overall == "PARTIAL" else C.FAIL)
    print(f"\n  {oc}{C.BOLD}Overall: {overall}  ({passed}/{total} passed){C.END}\n", flush=True)

    # Parsed by Jenkins pipeline to get status + file paths
    print(f"OUTPUT_TXT={txt_path}",    flush=True)
    print(f"OUTPUT_CSV={csv_path}",    flush=True)
    print(f"OUTPUT_JSON={json_path}",  flush=True)
    print(f"OUTPUT_XLSX={xlsx_path}",  flush=True)
    print(f"OVERALL_STATUS={overall}", flush=True)

    # Fail the pytest build if every channel failed
    assert overall != "FAIL", \
        f"PTMP Channel Test FAILED — {total - passed}/{total} channels failed"