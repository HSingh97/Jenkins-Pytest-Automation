import re
import time
import platform
import subprocess
import os
import paramiko
import csv
from datetime import datetime

# Try importing openpyxl for Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font

    excel_available = True
except ImportError:
    print("!!! WARNING: 'openpyxl' module not found. Excel report will NOT be generated.")
    excel_available = False

# CONFIGURATION
file_name_base = 'Nov_27_PTMP_Test'
txt_file_name = f'{file_name_base}.txt'
csv_file_name = f'{file_name_base}.csv'
xlsx_file_name = f'{file_name_base}_Summary.xlsx'

local_ip = "192.168.1.26"  # BSU IP

# Remote SUs
remote_ips = [
    "192.168.1.28", "192.168.1.29", "192.168.1.30", "192.168.1.31",
    "192.168.1.32", "192.168.1.33", "192.168.1.34", "192.168.1.35"
]

custom_channel_list = ["160", "161"]

bandwidth = ["HT80"]
countries = ["5GHz"]
radio = "radio1"

username_SSH = "root"
passwords_ssh = "admin"

channel_results = []

# ANSI Colors
class Colors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'

if radio == "radio1":
    radio_index = '1'
    radio_oid = '2'
elif radio == "radio2":
    radio_index = '2'
    radio_oid = '3'
else:
    print("Select a valid Radio")
    exit()

def init_csv():
    with open(csv_file_name, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(
            ["Timestamp", "Channel", "Node_IP", "Local_SNR_A1/A2", "Remote_SNR_A1/A2", "Tx/Rx", "Uptime_Raw", "Status",
             "Notes"])

def log_to_csv(data_row):
    with open(csv_file_name, mode='a', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(data_row)

def parse_uptime_to_seconds(uptime_str):
    try:
        if not uptime_str or uptime_str == "Err": return -1
        parts = [int(x) for x in uptime_str.split(':')]
        parts.reverse()
        total_seconds = 0
        if len(parts) > 0: total_seconds += parts[0]
        if len(parts) > 1: total_seconds += parts[1] * 60
        if len(parts) > 2: total_seconds += parts[2] * 3600
        if len(parts) > 3: total_seconds += parts[3] * 86400
        return total_seconds
    except:
        return -1

def start_test():
    init_csv()
    with open(txt_file_name, 'w') as f:
        f.write(f"TEST STARTED: {datetime.now()}\n")

    k = 0
    while k < len(countries):
        check_single_ip_reachability(local_ip)
        current_country = countries[k]
        print(f"{Colors.HEADER}Current Country : {current_country}{Colors.ENDC}")

        if current_country == "US 5GHz ALL":
            country_code = 5012
        elif current_country == "5GHz":
            country_code = 5019
        elif current_country == "Canada":
            country_code = 124
        elif current_country == "US 5GHz Non-DFS":
            country_code = 5011
        else:
            country_code = 5012

        with open(txt_file_name, 'a') as f:
            f.write(f"\n{'=' * 60}\n\t\tCountry : {current_country}\n{'=' * 60}\n")

        i = 0
        while i < len(bandwidth):
            check_single_ip_reachability(local_ip)
            current_bw = bandwidth[i]
            print(f"{Colors.HEADER}Current Bandwidth : {current_bw}{Colors.ENDC}")

            #set_oid_with_retry("Bandwidth", f".1.3.6.1.4.1.52619.1.1.1.1.1.7.{radio_oid}", "s", current_bw)
            check_single_ip_reachability(local_ip)

            # --- CHANNEL SELECTION LOGIC ---
            if len(custom_channel_list) > 0:
                print(f"{Colors.OKBLUE}Using Custom Channel List: {custom_channel_list}{Colors.ENDC}")
                channel_list = custom_channel_list
            else:
                channel_list = get_channel_list(local_ip, country_code, current_bw)

            j = 0
            while j < len(channel_list):
                current_chan = channel_list[j]
                print(f"{Colors.OKBLUE}--> Testing Channel : {current_chan}{Colors.ENDC}")

                set_oid_with_retry("Channel", f".1.3.6.1.4.1.52619.1.1.1.1.1.9.{radio_oid}", "i", current_chan)

                print("Waiting for SUs to connect...")
                links_up = wait_for_connection(remote_ips)

                # CHECK 1: IF LINKS DO NOT FORM, SKIP CHANNEL
                if not links_up:
                    print(
                        f"{Colors.FAIL}Link Timeout! Not all SUs connected. Skipping Channel {current_chan}{Colors.ENDC}")

                    # Log Failure
                    log_to_csv(
                        [datetime.now().strftime("%H:%M:%S"), current_chan, "ALL", "-/-", "-/-", "-/-", "-", "FAIL",
                         "Link Timeout"])
                    channel_results.append({'Channel': current_chan, 'Status': 'FAIL', 'Notes': 'Link Timeout'})

                    with open(txt_file_name, 'a') as f:
                        f.write(f"\n[SUMMARY] Channel {current_chan}: FAILED (Link Timeout)\n")

                    # IMMEDIATE SKIP
                    j += 1
                    continue

                print(f"{Colors.OKGREEN}All Links Formed in Channel {current_chan}{Colors.ENDC}")
                print(f"--> Starting 5-minute stability check for Channel {current_chan}...")

                monitor_duration = 120
                interval = 5
                elapsed_time = 0

                incidents = []
                previous_uptime_map = {}

                # MONITORING LOOP
                while elapsed_time < monitor_duration:
                    loop_start = time.time()
                    timestamp = datetime.now().strftime("%H:%M:%S")

                    iteration_issues = get_and_log_linkstats(current_chan, timestamp, previous_uptime_map)

                    zero_snr_detected = False

                    if iteration_issues:
                        incidents.extend(iteration_issues)
                        # Check if any issue is a Zero SNR
                        for issue in iteration_issues:
                            if "Zero SNR" in issue:
                                zero_snr_detected = True

                    # CHECK 2: IMMEDIATE EXIT ON ZERO SNR
                    if zero_snr_detected:
                        print(
                            f"{Colors.FAIL}!!! CRITICAL FAILURE: Zero SNR Detected. Aborting Channel {current_chan} !!!{Colors.ENDC}")
                        break  # Break the while loop immediately

                    process_time = time.time() - loop_start
                    sleep_time = max(0, interval - process_time)
                    try:
                        time.sleep(sleep_time)
                    except KeyboardInterrupt:
                        print("\nTest Stopped by User. Generating Report...")
                        generate_excel_report()
                        exit()

                    elapsed_time += interval
                    print(f"Monitoring... {elapsed_time}/{monitor_duration}s")

                # --- END OF MONITORING SUMMARY ---
                unique_incidents = list(set(incidents))

                if len(incidents) > 0:
                    status = "FAIL"
                    notes = "; ".join(unique_incidents[:3])
                    if len(unique_incidents) > 3: notes += "..."
                else:
                    status = "PASS"
                    notes = "Stable"

                channel_results.append({
                    'Channel': current_chan,
                    'Status': status,
                    'Notes': notes
                })

                with open(txt_file_name, 'a') as f:
                    if status == "FAIL":
                        f.write(f"\n[SUMMARY] Issues on Channel {current_chan}:\n")
                        for inc in incidents: f.write(f" - {inc}\n")
                    else:
                        f.write(f"\n[SUMMARY] Channel {current_chan}: STABLE\n")

                j += 1
            i += 1
        k += 1

    generate_excel_report()

def generate_excel_report():
    if not excel_available: return
    print(f"\n{Colors.HEADER}Generating Excel Report: {xlsx_file_name}...{Colors.ENDC}")
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Summary"
    headers = ["Channel", "Status", "Details/Notes"]
    ws.append(headers)
    fill_pass = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    fill_fail = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    font_bold = Font(bold=True)
    for cell in ws[1]: cell.font = font_bold

    for row_idx, data in enumerate(channel_results, start=2):
        ws.cell(row=row_idx, column=1, value=data['Channel'])
        cell_status = ws.cell(row=row_idx, column=2, value=data['Status'])
        ws.cell(row=row_idx, column=3, value=data['Notes'])
        if data['Status'] == "PASS":
            cell_status.fill = fill_pass
        else:
            cell_status.fill = fill_fail

    wb.save(xlsx_file_name)
    print(f"{Colors.OKGREEN}Report Saved Successfully!{Colors.ENDC}")

def check_single_ip_reachability(ip):
    t = 0
    while True:
        if t >= 200: return False
        if ping(ip): return True
        time.sleep(10);
        t += 10


def wait_for_connection(ip_list):
    timeout = 0
    # Wait up to 120s for connection
    while timeout < 120:
        all_up = True
        for ip in ip_list:
            if not ping(ip):
                all_up = False;
                break
        if all_up: return True
        time.sleep(5);
        timeout += 5
    return False


def set_oid_with_retry(name, oid, type_char, value):
    attempts = 0
    while attempts < 3:
        cmd = f"snmpset -v 2c -c private {local_ip} {oid} {type_char} {value}"
        apply_cmd = f"snmpset -v 2c -c private {local_ip} .1.3.6.1.4.1.52619.1.2.1.1.0 i 1"
        try:
            subprocess.check_output(cmd, shell=True, stderr=subprocess.STDOUT)
            time.sleep(2)
            subprocess.check_output(apply_cmd, shell=True, stderr=subprocess.STDOUT)
            if name == "Bandwidth":
                time.sleep(60)
            else:
                time.sleep(30)
            return True
        except subprocess.CalledProcessError as e:
            attempts += 1
            print(f"{Colors.FAIL}Error setting {name} (Attempt {attempts}/3): {e.output.decode().strip()}{Colors.ENDC}")
            time.sleep(5)
    print(f"{Colors.FAIL}!!! FAILED TO SET {name} AFTER 3 ATTEMPTS !!!{Colors.ENDC}")
    return False

def get_and_log_linkstats(channel, timestamp, previous_uptime_map):
    issues = []
    found_data = {}

    i = 1
    while i < 33:
        try:
            cmd_ip = f"snmpget -v 2c -c private {local_ip} .1.3.6.1.4.1.52619.1.3.3.1.4.{radio_oid}.{i}"
            out_ip = subprocess.check_output(cmd_ip, shell=True).decode("utf-8")
        except:
            i += 1; continue

        if "No Such Instance" in out_ip: i += 1; continue
        match_ip = re.search(r'IpAddress:\s*([\d.]+)', out_ip)
        ip_addr = match_ip.group(1) if match_ip else "-"
        if ip_addr == "-" or ip_addr == "0.0.0.0": i += 1; continue

        def get_val(oid_suffix):
            try:
                c = f"snmpget -v 2c -c private {local_ip} .1.3.6.1.4.1.52619.1.3.3.1.{oid_suffix}.{radio_oid}.{i}"
                o = subprocess.check_output(c, shell=True).decode("utf-8")
                m = re.search(r'INTEGER:\s*(\d+)', o)
                return m.group(1) if m else "-"
            except:
                return "-"

        def get_uptime_raw():
            try:
                c = f"snmpget -v 2c -c private {local_ip} .1.3.6.1.4.1.52619.1.3.3.1.52.{radio_oid}.{i}"
                o = subprocess.check_output(c, shell=True).decode("utf-8")
                raw_val = o.split('=')[-1].strip()
                raw_val = re.sub(r'^(STRING|Timeticks|INTEGER):\s*', '', raw_val)
                raw_val = raw_val.replace('"', '').replace('(', '').replace(')', '').strip()
                return raw_val
            except:
                return "Err"

        found_data[ip_addr] = {
            'l_a1': get_val("13"), 'l_a2': get_val("14"),
            'r_a1': get_val("15"), 'r_a2': get_val("16"),
            'tx': get_val("10"), 'rx': get_val("9"),
            'uptime': get_uptime_raw()
        }
        i += 1

    print(f"\n{Colors.BOLD}{'IP':<16} | {'L_SNR':<7} | {'R_SNR':<7} | {'UPTIME':<12} | {'STATUS'}{Colors.ENDC}")

    for expected_ip in remote_ips:
        status = "PASS"
        note = ""
        row_color = Colors.OKGREEN

        if expected_ip in found_data:
            d = found_data[expected_ip]
            uptime_str = d['uptime']

            # Link Reformed Check
            current_seconds = parse_uptime_to_seconds(uptime_str)
            prev_seconds = previous_uptime_map.get(expected_ip, -1)

            if prev_seconds != -1 and current_seconds != -1:
                if current_seconds < prev_seconds:
                    status = "FAIL"
                    note = "LINK REFORMED"
                    row_color = Colors.FAIL
                    issues.append(f"Node {expected_ip}: LINK REFORMED")

            if current_seconds != -1: previous_uptime_map[expected_ip] = current_seconds

            # Zero SNR Check
            if (d['l_a1'] == "0" or d['l_a2'] == "0" or d['r_a1'] == "0" or d['r_a2'] == "0"):
                status = "FAIL"
                note += " Zero SNR" if note == "" else " & Zero SNR"
                row_color = Colors.FAIL
                issues.append(f"Node {expected_ip}: Zero SNR")

            print(
                f"{row_color}{expected_ip:<16} | {d['l_a1']}/{d['l_a2']:<5} | {d['r_a1']}/{d['r_a2']:<5} | {uptime_str:<12} | {status} {note}{Colors.ENDC}")

            log_to_csv([timestamp, channel, expected_ip, f"{d['l_a1']}/{d['l_a2']}", f"{d['r_a1']}/{d['r_a2']}",
                        f"{d['tx']}/{d['rx']}", uptime_str, status, note])
            with open(txt_file_name, 'a') as f:
                f.write(
                    f"[{timestamp}] IP:{expected_ip}\tL:{d['l_a1']}/{d['l_a2']}\tR:{d['r_a1']}/{d['r_a2']}\tUp:{uptime_str}\t{status} {note}\n")

        else:
            status = "MISSING"
            note = "Not in BSU Table"
            row_color = Colors.FAIL
            issues.append(f"Node {expected_ip}: MISSING from SNMP Table")
            print(f"{row_color}{expected_ip:<16} | {'-/-':<7} | {'-/-':<7} | {'-':<12} | {status}{Colors.ENDC}")
            log_to_csv([timestamp, channel, expected_ip, "-/-", "-/-", "-/-", "-", "FAIL", "Missing in SNMP"])

    return issues

def get_channel_list(ip, country, bandwidth):
    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(ip, username=username_SSH, password=passwords_ssh)
        cmd = f'/usr/sbin/kwn_get_supp_chan.sh {radio_index} {country} {bandwidth}'
        _, stdout, _ = ssh.exec_command(cmd)
        nums = re.findall(r'\b\d+\b', stdout.read().decode('utf-8'))
        ssh.close()
        return nums[::2]
    except:
        return []

def ping(host):
    param = '-n' if platform.system().lower() == 'windows' else '-c'
    return subprocess.call(['ping', param, '1', host], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL) == 0


if __name__ == "__main__":
    start_test()